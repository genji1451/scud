
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import datetime
import locale
from pathlib import Path
import sys

try:
    locale.setlocale(locale.LC_TIME, 'ru_RU.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_TIME, 'ru_RU')
    except:
        pass

DEFAULT_INPUT_FILES = [
    'nov-feb 11.xlsx',
    'nov-feb 11 2.xlsx',   
    'nov-feb 11 3.xlsx',

]
OUTPUT_FILE = 'weekly_report.xlsx'
BREAKS_FILE = 'breaks_report.xlsx'

EXCLUDE_NAMES = [
    'Сапаев Дмитрий',
    'Чамурлиева Майя',
    'Калинкин Илья',
    'Голубитченко',
    'Мажникова',
    'Петрухин',
    'Сапаева Мария',
    'Сапаева Светлана',
    'Свободная 2',
    'Свободная 8',  # Это тот же самый 'Все свободные'
    'Сусоева',
    'Ястребова',
]

# Явное переименование некоторых проходов «Свободная N» в реальные ФИО
NAME_REMAP = {
    'Свободная 11': 'Крафт Сергей Алексеевич',
    'Свободная 5': 'Алхимова Ирина Алексеевна',
    'Свободная 7': 'Одиноков Сергей Владимирович',
    'Свободная 9': 'Хамищев Роман Андреевич',
    'Свободная 3': 'Стяжков Дмитрий Алексеевич',
    'Чернова Ирина Валентиновна': 'Чернова Елена Валентиновна',
}

# Строки для исключения из статистики: (Сотрудник, Дата). Эти дни не попадут в отчёты и на сайт.
EXCLUDE_ROWS = [
    # ('Фамилия Имя Отчество', '03.12.2025'),
    # ('Другой сотрудник', '15.01.2026'),
]

def read_input_files(input_files: list[str]) -> pd.DataFrame:
    dfs: list[pd.DataFrame] = []
    for file in input_files:
        path = Path(file)
        if not path.exists():
            print(f"WARNING: File not found: {file}")
            continue
        print(f"Reading {file}...")
        try:
            df_part = pd.read_excel(file, header=3)
        except Exception as e:
            print(f"Error reading file {file}: {e}")
            continue
        df_part['__source_file'] = str(file)
        dfs.append(df_part)

    if not dfs:
        raise FileNotFoundError(
            "No input Excel files were loaded. "
            "Pass file paths as arguments, e.g.: "
            "python3 generate_report.py \"nov-feb 11.xlsx\" \"nov-feb 11 2.xlsx\""
        )

    df_all = pd.concat(dfs, ignore_index=True)

    # Удаляем возможные дубли событий при склейке файлов
    dedupe_cols = [
        c for c in ['Фамилия', 'Имя', 'Отчество', 'Дата события', 'Устройство', 'Событие', 'Выход']
        if c in df_all.columns
    ]
    if dedupe_cols:
        before = len(df_all)
        df_all = df_all.drop_duplicates(subset=dedupe_cols)
        after = len(df_all)
        if after != before:
            print(f"Deduplicated events: {before - after}")

    return df_all

def generate_report():
    try:
        input_files = sys.argv[1:] if len(sys.argv) > 1 else DEFAULT_INPUT_FILES
        df = read_input_files(input_files)
    except Exception as e:
        print(f"Error reading input files: {e}")
        return

    # Basic cleanup
    df = df.dropna(subset=['Фамилия'])
    df = df[df['Фамилия'] != 'Охрана']
    
    # Filter relevant events
    valid_events = ['Проход по идентификатору', 'Проход по команде от ДУ']
    df = df[df['Событие'].isin(valid_events)]
    
    # Create Full Name
    df['Full Name'] = df['Фамилия'].astype(str) + ' ' + \
                      df['Имя'].astype(str).replace('nan', '') + ' ' + \
                      df['Отчество'].astype(str).replace('nan', '')
    df['Full Name'] = df['Full Name'].str.strip()
    df['Full Name'] = df['Full Name'].replace(NAME_REMAP)
    
    # Apply Blacklist
    original_count = len(df)
    for excluded in EXCLUDE_NAMES:
        df = df[~df['Full Name'].astype(str).str.contains(excluded, case=False)]
    
    print(f"Filtered out {original_count - len(df)} events based on exclusion list.")
    
    # Parse Date
    df['Datetime'] = pd.to_datetime(df['Дата события'], errors='coerce')
    df = df.dropna(subset=['Datetime'])
    df['Date'] = df['Datetime'].dt.date
    df['Time'] = df['Datetime'].dt.time

    # --- Фильтрация периода и выходных ---
    # Удаляем весь ноябрь (месяц == 11) и все выходные (суббота/воскресенье)
    df = df[df['Datetime'].dt.month != 11]
    df = df[df['Datetime'].dt.weekday < 5]
    
    # Determine Direction
    def get_direction(row):
        if row['Выход'] == 'Офисное здание':
            return 'In'
        elif row['Выход'] == 'Неконтролируемая территория':
            return 'Out'
        return 'Unknown'
        
    df['Direction'] = df.apply(get_direction, axis=1)
    
    # Sort for processing
    df = df.sort_values(by=['Full Name', 'Datetime'])

    all_dates = sorted(df['Date'].unique())
    
    # --- Process Data ---
    
    detailed_rows = []
    # per-day per-employee summary (for Excel + сайт)
    breaks_summary_rows = [] 
    # детализация каждого отдельного выхода/возврата
    detailed_breaks_rows = [] 
    # данные для табеля (первый вход / последний выход)
    employee_data = {} 
    
    grouped = df.groupby(['Full Name', 'Date'])
    exclude_set = {(e, d) for e, d in EXCLUDE_ROWS}
    if exclude_set:
        print(f"Исключено строк из статистики: {len(exclude_set)} (см. EXCLUDE_ROWS)")

    for (name, date), group in grouped:
        date_str = date.strftime('%d.%m.%Y')
        if (name, date_str) in exclude_set:
            continue
        if name not in employee_data:
            employee_data[name] = {}
            
        day_events = group.sort_values('Datetime').to_dict('records')
        
        # --- Summary Stats Calculation ---
        entries = [e for e in day_events if e['Direction'] == 'In']
        exits = [e for e in day_events if e['Direction'] == 'Out']
        
        first_in = entries[0]['Datetime'] if entries else None
        final_out_time = None
        
        if first_in:
            valid_exits = [e for e in exits if e['Datetime'] > first_in]
            if valid_exits:
                final_out_time = valid_exits[-1]['Datetime']
            else:
                final_out_time = None
        else:
            if exits:
                final_out_time = exits[-1]['Datetime']

        employee_data[name][date] = {
            'in': first_in,
            'out': final_out_time
        }
        
        # --- Breaks Calculation ---
        day_break_count = 0
        day_break_seconds = 0
        lunch_seconds = 0  # время на обед
        smoke_seconds = 0  # время на перекуры
        day_breaks_list = []  # список всех перерывов за день для JSON
        
        for i, event in enumerate(day_events):
            duration_str = ""
            
            if event['Direction'] == 'Out':
                next_in = None
                for j in range(i + 1, len(day_events)):
                    if day_events[j]['Direction'] == 'In':
                        next_in = day_events[j]
                        break
                
                if next_in:
                    diff = next_in['Datetime'] - event['Datetime']
                    total_seconds = int(diff.total_seconds())
                    
                    day_break_count += 1
                    day_break_seconds += total_seconds
                    
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    duration_str = f"{hours}ч {minutes}м" if hours > 0 else f"{minutes}м"
                    
                    # Классификация перерыва: >= 30 минут = обед, < 30 минут = перекур
                    break_type = 'Обед' if total_seconds >= 1800 else 'Перекур'
                    
                    # Подсчет времени по типам перерывов
                    if break_type == 'Обед':
                        lunch_seconds += total_seconds
                    else:
                        smoke_seconds += total_seconds
                    
                    detailed_breaks_rows.append({
                        'Сотрудник': name,
                        'Дата': date.strftime('%d.%m.%Y'),
                        'Время выхода': event['Time'].strftime('%H:%M:%S'),
                        'Время возвращения': next_in['Time'].strftime('%H:%M:%S'),
                        'Длительность': duration_str,
                        'Тип': break_type
                    })
                    
                    # Сохраняем для JSON сайта
                    day_breaks_list.append({
                        'Время выхода': event['Time'].strftime('%H:%M'),
                        'Время возвращения': next_in['Time'].strftime('%H:%M'),
                        'Длительность_сек': total_seconds,
                        'Тип': break_type
                    })

            detailed_rows.append({
                'Сотрудник': name,
                'Дата': date.strftime('%d.%m.%Y'),
                'Время': event['Time'],
                'Направление': 'Вход' if event['Direction'] == 'In' else ('Выход' if event['Direction'] == 'Out' else 'Неизвестно'),
                'Длительность отсутствия': duration_str,
                'Событие': event['Событие'],
                'Турникет/Дверь': event['Устройство']
            })
            
        b_hours = day_break_seconds // 3600
        b_minutes = (day_break_seconds % 3600) // 60
        total_duration_str = f"{b_hours}ч {b_minutes}м" if (b_hours > 0 or b_minutes > 0) else "-"
        
        # --- Work Time Calculation (с учетом перерывов) ---
        work_time_str = "-"
        net_seconds_value = None
        net_minus_lunch_seconds = None  # время минус обед
        net_minus_smoke_seconds = None  # время минус перекуры
        first_in_str = first_in.strftime('%H:%M') if first_in else ""
        last_out_str = final_out_time.strftime('%H:%M') if final_out_time else ""

        if first_in and final_out_time:
            gross_seconds = int((final_out_time - first_in).total_seconds())
            net_seconds = gross_seconds - day_break_seconds
            net_seconds_value = max(net_seconds, 0)
            
            # Время работы минус обед (но с перекурами)
            net_minus_lunch_seconds = max(gross_seconds - lunch_seconds, 0)
            
            # Время работы минус перекуры (но с обедом)
            net_minus_smoke_seconds = max(gross_seconds - smoke_seconds, 0)
            
            if net_seconds_value > 0:
                w_hours = net_seconds_value // 3600
                w_minutes = (net_seconds_value % 3600) // 60
                work_time_str = f"{w_hours}ч {w_minutes}м"
            else:
                work_time_str = "0ч 0м"
        else:
            net_seconds_value = 0
            net_minus_lunch_seconds = 0
            net_minus_smoke_seconds = 0
        
        breaks_summary_rows.append({
            'Сотрудник': name,
            'Дата': date.strftime('%d.%m.%Y'),
            'Первый вход': first_in_str,
            'Последний выход': last_out_str,
            'Количество выходов': day_break_count if day_break_count > 0 else "-",
            'Общее время отсутствия': total_duration_str,
            'Чистое рабочее время': work_time_str,
            # для сайта: чистое рабочее время в секундах
            'net_seconds': int(net_seconds_value) if net_seconds_value is not None else 0,
            'net_minus_lunch_seconds': int(net_minus_lunch_seconds) if net_minus_lunch_seconds is not None else 0,
            'net_minus_smoke_seconds': int(net_minus_smoke_seconds) if net_minus_smoke_seconds is not None else 0,
            'lunch_seconds': int(lunch_seconds),
            'smoke_seconds': int(smoke_seconds),
            # список перерывов для сайта (JSON-совместимый формат)
            'breaks': day_breaks_list if day_breaks_list else []
        })

    # --- Create DataFrames ---
    
    rows = []
    employees = sorted(employee_data.keys())
    
    for emp in employees:
        row = {'Сотрудник': emp}
        for d in all_dates:
            data = employee_data[emp].get(d, {})
            t_in = data.get('in')
            t_out = data.get('out')
            
            str_in = t_in.strftime('%H:%M') if t_in else ''
            str_out = t_out.strftime('%H:%M') if t_out else ''
            
            row[f"{d}_In"] = str_in
            row[f"{d}_Out"] = str_out
            
        rows.append(row)
        
    summary_df = pd.DataFrame(rows)
    
    # полная сводка по перерывам (включая сырые секунды)
    breaks_summary_df_all = pd.DataFrame(breaks_summary_rows)
    breaks_summary_df = breaks_summary_df_all.copy()
    if not breaks_summary_df.empty:
        breaks_summary_df = breaks_summary_df[['Сотрудник', 'Дата', 'Первый вход', 'Последний выход', 'Количество выходов', 'Общее время отсутствия', 'Чистое рабочее время']]
        breaks_summary_df = breaks_summary_df.sort_values(by=['Сотрудник', 'Дата'])
    
    detailed_df = pd.DataFrame(detailed_rows)
    if not detailed_df.empty:
        detailed_df = detailed_df[['Сотрудник', 'Дата', 'Время', 'Направление', 'Длительность отсутствия', 'Событие', 'Турникет/Дверь']]

    detailed_breaks_df = pd.DataFrame(detailed_breaks_rows)
    if not detailed_breaks_df.empty:
        detailed_breaks_df = detailed_breaks_df[['Сотрудник', 'Дата', 'Время выхода', 'Время возвращения', 'Длительность']]
        detailed_breaks_df = detailed_breaks_df.sort_values(by=['Сотрудник', 'Дата', 'Время выхода'])

    # --- Export JSON for веб-сайт ---
    # Экспортируем данные с перерывами в JSON
    import json
    site_data = []
    for _, row in breaks_summary_df_all.iterrows():
        site_data.append({
            'Сотрудник': row['Сотрудник'],
            'Дата': row['Дата'],
            'Первый вход': row['Первый вход'],
            'Последний выход': row['Последний выход'],
            'net_seconds': int(row['net_seconds']),  # чистое время (минус все перерывы)
            'work_hours': row['net_seconds'] / 3600.0,
            'net_minus_lunch_seconds': int(row.get('net_minus_lunch_seconds', 0)),  # минус обед
            'work_minus_lunch_hours': row.get('net_minus_lunch_seconds', 0) / 3600.0,
            'net_minus_smoke_seconds': int(row.get('net_minus_smoke_seconds', 0)),  # минус перекуры
            'work_minus_smoke_hours': row.get('net_minus_smoke_seconds', 0) / 3600.0,
            'lunch_seconds': int(row.get('lunch_seconds', 0)),
            'smoke_seconds': int(row.get('smoke_seconds', 0)),
            'breaks': row.get('breaks', [])
        })
    
    with open('work_summary.json', 'w', encoding='utf-8') as f:
        json.dump(site_data, f, ensure_ascii=False, indent=2)

    # --- Write Main Report ---
    print(f"Writing to {OUTPUT_FILE}...")
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='Табель', index=False)
        breaks_summary_df.to_excel(writer, sheet_name='Сводка по перерывам', index=False)
        detailed_df.to_excel(writer, sheet_name='Детальный журнал', index=False)
        
    format_workbook(OUTPUT_FILE, all_dates)
    
    # --- Write Separate Breaks Report ---
    print(f"Writing to {BREAKS_FILE}...")
    with pd.ExcelWriter(BREAKS_FILE, engine='openpyxl') as writer:
        breaks_summary_df.to_excel(writer, sheet_name='Итоговая сводка', index=False)
        detailed_breaks_df.to_excel(writer, sheet_name='Детальные перерывы', index=False)
    
    format_simple_workbook(BREAKS_FILE)
    
    print("Done!")

def format_workbook(filename, all_dates):
    wb = load_workbook(filename)
    
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
    fill_weekend = PatternFill(start_color='FFEEEE', end_color='FFEEEE', fill_type='solid')

    ws = wb['Табель']
    ws.insert_rows(1)
    ws.merge_cells('A1:A2')
    ws['A1'] = 'Сотрудник'
    ws['A1'].alignment = center_align
    ws['A1'].font = header_font
    ws['A1'].border = border_style
    ws['A2'].border = border_style

    col_idx = 2
    for d in all_dates:
        col_letter_in = get_column_letter(col_idx)
        col_letter_out = get_column_letter(col_idx + 1)
        day_str = d.strftime('%d.%m (%a)')
        
        ws.merge_cells(f'{col_letter_in}1:{col_letter_out}1')
        cell_date = ws[f'{col_letter_in}1']
        cell_date.value = day_str
        cell_date.alignment = center_align
        cell_date.font = header_font
        cell_date.border = border_style
        ws[f'{col_letter_out}1'].border = border_style
        
        cell_in = ws[f'{col_letter_in}2']
        cell_in.value = "Пришел"
        cell_in.alignment = center_align
        cell_in.font = header_font
        cell_in.border = border_style
        
        cell_out = ws[f'{col_letter_out}2']
        cell_out.value = "Ушел"
        cell_out.alignment = center_align
        cell_out.font = header_font
        cell_out.border = border_style
        
        if d.weekday() >= 5:
             cell_date.fill = fill_weekend
             cell_in.fill = fill_weekend
             cell_out.fill = fill_weekend
             
        col_idx += 2
        
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.border = border_style
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.column == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    ws.column_dimensions['A'].width = 30
    for i in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 10

    for sheet_name in wb.sheetnames:
        if sheet_name == 'Табель': continue
        ws_sheet = wb[sheet_name]
        for row in ws_sheet.iter_rows():
            for cell in row:
                cell.border = border_style
                if cell.column == 1: 
                     cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                     cell.alignment = Alignment(horizontal='center', vertical='center')
                
                if row[0].row == 1:
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for col in ws_sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                 try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                 except: pass
            ws_sheet.column_dimensions[col_letter].width = max_len + 3

    wb.save(filename)

def format_simple_workbook(filename):
    wb = load_workbook(filename)
    header_font = Font(bold=True)
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border_style
                if cell.column == 1:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                if row[0].row == 1:
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                 try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                 except: pass
            ws.column_dimensions[col_letter].width = max_len + 3
            
    wb.save(filename)

if __name__ == "__main__":
    generate_report()
